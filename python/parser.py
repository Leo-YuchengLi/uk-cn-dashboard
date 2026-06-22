#!/usr/bin/env python3
"""
UK-CN Weekly Report Excel Parser
Parses BSP ticket data Excel → SQLite fact tables
"""

import sys
import os
import re
import json
import sqlite3
import time
from pathlib import Path
from datetime import datetime

import pandas as pd
import openpyxl


DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'airline.db')


# ─── Auxiliary sheet readers ───

def read_tmc_iata(wb):
    """Read TMC-IATA sheet → {agency_no_str: tmc_name}"""
    ws = wb['TMC - IATA']
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    tmc_map = {}
    for col_idx in range(2, len(header)):
        tmc_name = header[col_idx]
        if not tmc_name:
            continue
        for row in rows[1:]:
            if col_idx < len(row) and row[col_idx] is not None:
                val = str(row[col_idx]).strip().replace('\xa0', '')
                if val:
                    tmc_map[val] = str(tmc_name)
    return tmc_map


def read_trip_iata(wb):
    """Read Trip IATA sheet → set of agency_no_str"""
    ws = wb['Trip IATA']
    trip_set = set()
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None:
            trip_set.add(str(row[0]).strip())
    return trip_set


def read_ota_agencies(wb):
    """Read OTA sheet → set of agency_no_str"""
    ws = wb['OTA']
    agencies = set()
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                v = str(val).strip()
                if v.isdigit() and len(v) >= 7:
                    agencies.add(v)
    return agencies


def read_consol_agencies(wb):
    """Read CONSOL sheet → set of agency_no_str"""
    ws = wb['CONSOL']
    agencies = set()
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if val is not None:
                v = str(val).strip()
                if v.isdigit() and len(v) >= 7:
                    agencies.add(v)
    return agencies


# ─── Sheet identification ───

def find_data_sheet(wb):
    """Find the main data sheet by column signature (not by name)."""
    required_cols = {'Trip Origin City', 'Trip Destination City', 'Pax',
                     'Dominant Operating Airline', 'Origin', 'Destination'}
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = {str(h) for h in row if h is not None}
            if required_cols.issubset(headers):
                return name
    raise ValueError("Cannot find data sheet with required columns")


def extract_snapshot_date(filepath):
    """Extract YYYYMMDD from filename."""
    basename = Path(filepath).stem
    match = re.search(r'(\d{8})', basename)
    if match:
        return match.group(1)
    raise ValueError(f"Cannot extract date from filename: {basename}")


# ─── Channel classification ───

def classify_channel(agency_no_str, source, tmc_map, trip_set, ota_set, consol_set):
    """
    Returns (channel, tmc_name)
    channel: 直销 / TMC / Trip.com / OTA / CONSOL / 分销
    """
    if source == 'Direct Contributed':
        return '直销', None

    if agency_no_str in trip_set:
        return 'Trip.com', None

    if agency_no_str in tmc_map:
        return 'TMC', tmc_map[agency_no_str]

    if agency_no_str in ota_set:
        return 'OTA', None

    if agency_no_str in consol_set:
        return 'CONSOL', None

    return '分销', None


# ─── Derived column computation ───

def compute_trip_type(origin, destination, trip_origin_country, trip_dest_country):
    """航程类型: 点点 / D+I / I+I"""
    od = f"{origin}{destination}"
    # segment OD == OD means 点点 (direct point-to-point)
    # Actually the formula checks if OD == SegOD, but we compute from origin/dest
    # 点点 = Origin+Destination == SegOrigin+SegDest (i.e., no connection)
    # For simplicity: we use the trip origin/dest country logic
    # D+I = one end is UK, other is China (direct international)
    # I+I = neither end is UK-China direct
    # 点点 = OD == SegOD
    # We'll compute this differently using the actual columns
    uk = 'United Kingdom'
    cn = "China (People's Republic of)"
    is_uk_cn = (
        (trip_origin_country == uk and trip_dest_country == cn) or
        (trip_origin_country == cn and trip_dest_country == uk)
    )
    if is_uk_cn:
        return 'D+I'
    else:
        return 'I+I'


def compute_cabin(cabin_class):
    """舱位: 两舱 / 经济舱"""
    if cabin_class and ('First' in str(cabin_class) or 'Business' in str(cabin_class)):
        return '两舱'
    return '经济舱'


def compute_owrt(ticket_type):
    """OW/RT: 单程 / 往返"""
    if ticket_type == 'One Way':
        return '单程'
    return '往返'


def compute_poo_direction(trip_origin_country):
    """POO境内外"""
    if trip_origin_country == "China (People's Republic of)":
        return '境内'
    return '境外'


def compute_zhifenxiao(source):
    """直分销"""
    if source == 'Direct Contributed':
        return '直销'
    return '分销'


# ─── City pair key ───

def make_citypair_key(city_a, city_b):
    """Undirected city pair key: alphabetical order"""
    cities = sorted([str(city_a), str(city_b)])
    return f"{cities[0]}<->{cities[1]}"


# ─── Database setup ───

def init_db(db_path):
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    c.executescript("""
    CREATE TABLE IF NOT EXISTS snapshots (
        snapshot_date TEXT PRIMARY KEY,
        file_name TEXT,
        trip_date_min TEXT,
        trip_date_max TEXT,
        total_pax INTEGER,
        loaded_at TEXT
    );

    CREATE TABLE IF NOT EXISTS fact_citypair (
        snapshot_date TEXT,
        origin_city TEXT,
        dest_city TEXT,
        citypair_key TEXT,
        airline TEXT,
        channel TEXT,
        trip_type TEXT,
        pax INTEGER,
        PRIMARY KEY (snapshot_date, origin_city, dest_city, airline, channel, trip_type)
    );

    CREATE TABLE IF NOT EXISTS fact_airportpair (
        snapshot_date TEXT,
        od TEXT,
        origin_apt TEXT,
        dest_apt TEXT,
        origin_city TEXT,
        dest_city TEXT,
        citypair_key TEXT,
        airline TEXT,
        channel TEXT,
        trip_type TEXT,
        cabin TEXT,
        owrt TEXT,
        pax INTEGER,
        PRIMARY KEY (snapshot_date, od, airline, channel, trip_type, cabin, owrt)
    );

    CREATE TABLE IF NOT EXISTS fact_airline (
        snapshot_date TEXT,
        airline TEXT,
        trip_type TEXT,
        cabin TEXT,
        pax INTEGER,
        PRIMARY KEY (snapshot_date, airline, trip_type, cabin)
    );

    CREATE TABLE IF NOT EXISTS fact_channel (
        snapshot_date TEXT,
        channel TEXT,
        agency_no TEXT,
        agency_name TEXT,
        tmc_name TEXT,
        airline TEXT,
        pax INTEGER,
        PRIMARY KEY (snapshot_date, channel, agency_no, airline)
    );

    CREATE TABLE IF NOT EXISTS fact_agency (
        snapshot_date TEXT,
        agency_no TEXT,
        agency_name TEXT,
        channel TEXT,
        tmc_name TEXT,
        airline TEXT,
        pax INTEGER,
        PRIMARY KEY (snapshot_date, agency_no, airline)
    );

    CREATE TABLE IF NOT EXISTS fact_month (
        snapshot_date TEXT,
        trip_month INTEGER,
        airline TEXT,
        pax INTEGER,
        PRIMARY KEY (snapshot_date, trip_month, airline)
    );

    CREATE TABLE IF NOT EXISTS dim_channel_map (
        agency_no TEXT PRIMARY KEY,
        channel TEXT,
        source TEXT,
        updated_at TEXT
    );

    -- Indexes for common queries
    CREATE INDEX IF NOT EXISTS idx_citypair_snap ON fact_citypair(snapshot_date);
    CREATE INDEX IF NOT EXISTS idx_citypair_key ON fact_citypair(citypair_key);
    CREATE INDEX IF NOT EXISTS idx_airportpair_snap ON fact_airportpair(snapshot_date);
    CREATE INDEX IF NOT EXISTS idx_airportpair_cpkey ON fact_airportpair(citypair_key);
    CREATE INDEX IF NOT EXISTS idx_channel_snap ON fact_channel(snapshot_date);
    CREATE INDEX IF NOT EXISTS idx_agency_snap ON fact_agency(snapshot_date);
    """)

    conn.commit()
    return conn


def seed_channel_map(conn, ota_set, consol_set):
    """Seed dim_channel_map with known OTA/CONSOL agencies."""
    c = conn.cursor()
    now = datetime.now().isoformat()
    for ag in ota_set:
        c.execute("""INSERT OR IGNORE INTO dim_channel_map (agency_no, channel, source, updated_at)
                     VALUES (?, 'OTA', 'excel_sheet', ?)""", (ag, now))
    for ag in consol_set:
        c.execute("""INSERT OR IGNORE INTO dim_channel_map (agency_no, channel, source, updated_at)
                     VALUES (?, 'CONSOL', 'excel_sheet', ?)""", (ag, now))
    conn.commit()


# ─── Main parsing logic ───

def parse_excel(filepath, db_path=None, uk_cn_only=True):
    """Main entry: parse Excel file into SQLite."""
    if db_path is None:
        db_path = DB_PATH

    filepath = os.path.expanduser(filepath)
    snapshot_date = extract_snapshot_date(filepath)

    print(f"[1/6] Opening workbook for auxiliary sheets...", flush=True)
    t0 = time.time()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Read auxiliary data
    tmc_map = read_tmc_iata(wb)
    trip_set = read_trip_iata(wb)
    ota_set = read_ota_agencies(wb)
    consol_set = read_consol_agencies(wb)

    # Also load dim_channel_map from DB for additional OTA/CONSOL
    db_path_abs = os.path.abspath(db_path)
    os.makedirs(os.path.dirname(db_path_abs), exist_ok=True)
    conn = init_db(db_path_abs)

    # Merge dim_channel_map into sets
    c = conn.cursor()
    for row in c.execute("SELECT agency_no, channel FROM dim_channel_map"):
        ag, ch = row
        if ch == 'OTA':
            ota_set.add(ag)
        elif ch == 'CONSOL':
            consol_set.add(ag)

    seed_channel_map(conn, ota_set, consol_set)

    # Find data sheet
    data_sheet = find_data_sheet(wb)
    wb.close()
    print(f"   Data sheet: '{data_sheet}' (aux loaded in {time.time()-t0:.1f}s)", flush=True)

    # Read main data with pandas (try calamine for speed, fallback to openpyxl)
    print(f"[2/6] Reading main data sheet with pandas...", flush=True)
    t1 = time.time()
    try:
        df = pd.read_excel(filepath, sheet_name=data_sheet, engine='calamine')
        print(f"   {len(df)} rows loaded via calamine in {time.time()-t1:.1f}s", flush=True)
    except Exception:
        df = pd.read_excel(filepath, sheet_name=data_sheet, engine='openpyxl')
        print(f"   {len(df)} rows loaded via openpyxl in {time.time()-t1:.1f}s", flush=True)

    # ─── Compute derived columns ───
    print(f"[3/6] Computing derived columns...", flush=True)
    t2 = time.time()

    # Agency number as string
    df['agency_no_str'] = df['Travel Agency Number'].fillna(0).astype(int).astype(str)

    # OD
    df['od_computed'] = df['Origin'].astype(str) + df['Destination'].astype(str)

    # Channel classification
    def row_channel(row):
        return classify_channel(
            row['agency_no_str'],
            row.get('Source', ''),
            tmc_map, trip_set, ota_set, consol_set
        )

    channels = df.apply(row_channel, axis=1, result_type='expand')
    df['channel'] = channels[0]
    df['tmc_name'] = channels[1]

    # Trip type: need to handle 点点 vs D+I vs I+I
    # 点点 = Origin+Destination == Segment Origin Airport + Segment Destination Airport
    df['_od'] = df['Origin'].astype(str) + df['Destination'].astype(str)
    df['_segod'] = df['Segment Origin Airport'].astype(str) + df['Segment Destination Airport'].astype(str)

    uk = 'United Kingdom'
    cn = "China (People's Republic of)"

    def compute_trip_type_row(row):
        if row['_od'] == row['_segod']:
            return '点点'
        is_uk_cn = (
            (row['Trip Origin Country Name'] == uk and row['Trip Destination Country Name'] == cn) or
            (row['Trip Origin Country Name'] == cn and row['Trip Destination Country Name'] == uk)
        )
        if is_uk_cn:
            return 'D+I'
        return 'I+I'

    df['trip_type'] = df.apply(compute_trip_type_row, axis=1)

    # Cabin
    df['cabin'] = df['O&D Dominant Cabin Class'].apply(
        lambda x: '两舱' if x and ('First' in str(x) or 'Business' in str(x)) else '经济舱'
    )

    # OW/RT
    df['owrt'] = df['Ticket Type'].apply(lambda x: '单程' if x == 'One Way' else '往返')

    # POO direction
    df['poo_direction'] = df['Trip Origin Country Name'].apply(
        lambda x: '境内' if x == cn else '境外'
    )

    # City pair key (undirected)
    df['citypair_key'] = df.apply(
        lambda r: make_citypair_key(r['Trip Origin City'], r['Trip Destination City']), axis=1
    )

    # UK-CN only filter
    if uk_cn_only:
        mask = (
            ((df['Trip Origin Country Name'] == uk) & (df['Trip Destination Country Name'] == cn)) |
            ((df['Trip Origin Country Name'] == cn) & (df['Trip Destination Country Name'] == uk))
        )
        df_filtered = df[mask].copy()
        print(f"   UK-CN filter: {len(df)} → {len(df_filtered)} rows", flush=True)
    else:
        df_filtered = df

    print(f"   Derived columns computed in {time.time()-t2:.1f}s", flush=True)

    # ─── Aggregate into fact tables ───
    print(f"[4/6] Aggregating fact tables...", flush=True)
    t3 = time.time()

    # Ensure Pax is numeric
    df_filtered['Pax'] = pd.to_numeric(df_filtered['Pax'], errors='coerce').fillna(0).astype(int)

    # Fill NaN to avoid groupby dropna issues
    df_filtered['Travel Agency Name'] = df_filtered['Travel Agency Name'].fillna('(直营)')
    df_filtered['tmc_name'] = df_filtered['tmc_name'].fillna('')

    # fact_citypair
    cp = df_filtered.groupby(
        ['Trip Origin City', 'Trip Destination City', 'citypair_key',
         'Dominant Operating Airline', 'channel', 'trip_type']
    )['Pax'].sum().reset_index()

    # fact_airportpair
    ap = df_filtered.groupby(
        ['od_computed', 'Origin', 'Destination',
         'Trip Origin City', 'Trip Destination City', 'citypair_key',
         'Dominant Operating Airline', 'channel', 'trip_type', 'cabin', 'owrt']
    )['Pax'].sum().reset_index()

    # fact_airline
    al = df_filtered.groupby(
        ['Dominant Operating Airline', 'trip_type', 'cabin']
    )['Pax'].sum().reset_index()

    # fact_channel: group by PK columns, keep first agency_name/tmc_name
    ch = df_filtered.groupby(
        ['channel', 'agency_no_str', 'Dominant Operating Airline']
    ).agg(
        pax=('Pax', 'sum'),
        agency_name=('Travel Agency Name', 'first'),
        tmc_name=('tmc_name', 'first'),
    ).reset_index()

    # fact_agency: group by PK columns
    ag = df_filtered.groupby(
        ['agency_no_str', 'Dominant Operating Airline']
    ).agg(
        pax=('Pax', 'sum'),
        agency_name=('Travel Agency Name', 'first'),
        channel=('channel', 'first'),
        tmc_name=('tmc_name', 'first'),
    ).reset_index()

    # fact_month
    mo = df_filtered.groupby(
        ['Trip Month', 'Dominant Operating Airline']
    )['Pax'].sum().reset_index()

    print(f"   Aggregated in {time.time()-t3:.1f}s", flush=True)

    # ─── Write to SQLite ───
    print(f"[5/6] Writing to SQLite...", flush=True)
    t4 = time.time()

    # Idempotent: delete existing data for this snapshot
    c = conn.cursor()
    for table in ['fact_citypair', 'fact_airportpair', 'fact_airline',
                   'fact_channel', 'fact_agency', 'fact_month', 'snapshots']:
        c.execute(f"DELETE FROM {table} WHERE snapshot_date = ?", (snapshot_date,))

    # snapshots
    trip_date_min = str(df_filtered['Trip Date'].min()) if 'Trip Date' in df_filtered else ''
    trip_date_max = str(df_filtered['Trip Date'].max()) if 'Trip Date' in df_filtered else ''
    total_pax = int(df_filtered['Pax'].sum())

    c.execute("""INSERT INTO snapshots VALUES (?,?,?,?,?,?)""",
              (snapshot_date, os.path.basename(filepath), trip_date_min, trip_date_max,
               total_pax, datetime.now().isoformat()))

    # fact_citypair
    cp_records = [(snapshot_date, r['Trip Origin City'], r['Trip Destination City'],
                   r['citypair_key'], r['Dominant Operating Airline'],
                   r['channel'], r['trip_type'], int(r['Pax']))
                  for _, r in cp.iterrows()]
    c.executemany("INSERT INTO fact_citypair VALUES (?,?,?,?,?,?,?,?)", cp_records)

    # fact_airportpair
    ap_records = [(snapshot_date, r['od_computed'], r['Origin'], r['Destination'],
                   r['Trip Origin City'], r['Trip Destination City'], r['citypair_key'],
                   r['Dominant Operating Airline'], r['channel'], r['trip_type'],
                   r['cabin'], r['owrt'], int(r['Pax']))
                  for _, r in ap.iterrows()]
    c.executemany("INSERT INTO fact_airportpair VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)", ap_records)

    # fact_airline
    al_records = [(snapshot_date, r['Dominant Operating Airline'], r['trip_type'],
                   r['cabin'], int(r['Pax']))
                  for _, r in al.iterrows()]
    c.executemany("INSERT INTO fact_airline VALUES (?,?,?,?,?)", al_records)

    # fact_channel
    ch_records = [(snapshot_date, r['channel'], r['agency_no_str'],
                   r['agency_name'], r['tmc_name'],
                   r['Dominant Operating Airline'], int(r['pax']))
                  for _, r in ch.iterrows()]
    c.executemany("INSERT INTO fact_channel VALUES (?,?,?,?,?,?,?)", ch_records)

    # fact_agency
    ag_records = [(snapshot_date, r['agency_no_str'], r['agency_name'],
                   r['channel'], r['tmc_name'],
                   r['Dominant Operating Airline'], int(r['pax']))
                  for _, r in ag.iterrows()]
    c.executemany("INSERT INTO fact_agency VALUES (?,?,?,?,?,?,?)", ag_records)

    # fact_month
    mo_records = [(snapshot_date, int(r['Trip Month']), r['Dominant Operating Airline'],
                   int(r['Pax']))
                  for _, r in mo.iterrows()]
    c.executemany("INSERT INTO fact_month VALUES (?,?,?,?)", mo_records)

    conn.commit()
    print(f"   Written to {db_path_abs} in {time.time()-t4:.1f}s", flush=True)

    # ─── Summary ───
    print(f"[6/6] Summary:", flush=True)
    print(f"   Snapshot: {snapshot_date}", flush=True)
    print(f"   Total Pax (UK-CN): {total_pax:,}", flush=True)
    print(f"   City pairs: {cp['citypair_key'].nunique()}", flush=True)
    print(f"   Airport pairs: {ap['od_computed'].nunique()}", flush=True)
    print(f"   Airlines: {al['Dominant Operating Airline'].nunique()}", flush=True)
    print(f"   Channels: {ch['channel'].nunique()}", flush=True)
    print(f"   Agencies: {ag['agency_no_str'].nunique()}", flush=True)
    print(f"   Total time: {time.time()-t0:.1f}s", flush=True)

    # Collect unclassified agencies for channel mapping maintenance
    unclassified = df_filtered[df_filtered['channel'] == '分销']
    if len(unclassified) > 0:
        unc_agencies = unclassified.groupby(['agency_no_str', 'Travel Agency Name'])['Pax'].sum()
        unc_agencies = unc_agencies.sort_values(ascending=False).head(20)
        print(f"\n   ⚠ Top unclassified 分销 agencies ({len(unclassified)} rows):", flush=True)
        for (ano, aname), pax in unc_agencies.items():
            print(f"     {ano} | {aname} | {pax:,} Pax", flush=True)

    conn.close()

    result = {
        'snapshot_date': snapshot_date,
        'total_pax': total_pax,
        'city_pairs': int(cp['citypair_key'].nunique()),
        'airport_pairs': int(ap['od_computed'].nunique()),
        'airlines': int(al['Dominant Operating Airline'].nunique()),
        'channels': int(ch['channel'].nunique()),
        'agencies': int(ag['agency_no_str'].nunique()),
    }
    print(json.dumps(result))
    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python parser.py <excel_file> [--all-markets] [--db DB_PATH]")
        sys.exit(1)

    filepath = sys.argv[1]
    uk_cn_only = '--all-markets' not in sys.argv
    db_path = None
    if '--db' in sys.argv:
        idx = sys.argv.index('--db')
        if idx + 1 < len(sys.argv):
            db_path = sys.argv[idx + 1]
    if not db_path:
        db_path = os.environ.get('DB_PATH', DB_PATH)
    parse_excel(filepath, db_path=db_path, uk_cn_only=uk_cn_only)
