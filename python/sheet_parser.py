#!/usr/bin/env python3
"""
Parse pre-computed pivot table data from each Excel sheet.
Stores structured data matching what the user sees in Excel.
"""

import sys
import os
import re
import json

# Fix Windows GBK encoding crash
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
import sqlite3
import time
from pathlib import Path

import openpyxl


DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'airline.db')
AIRLINES = ['CA', 'MU', 'CZ', 'HU', 'BA', 'HO', 'ZH', 'JD', 'GS', 'NZ']


def extract_snapshot_date(filepath):
    match = re.search(r'(\d{8})', Path(filepath).stem)
    if match:
        return match.group(1)
    raise ValueError(f"Cannot extract date from filename: {filepath}")


def read_sheet_as_grid(ws, max_row=120, max_col=80):
    """Read sheet into a 2D list for easier access."""
    grid = []
    for i, row in enumerate(ws.iter_rows(max_row=max_row, max_col=max_col, values_only=True)):
        grid.append(list(row))
        if i >= max_row:
            break
    return grid


# ─── Sheet 1: Month share - Pax ───

def parse_sheet1(wb):
    """Parse Month share - Pax sheet."""
    ws = wb['Month share - Pax']
    g = read_sheet_as_grid(ws)

    data = {}

    # Table 1: Airlines × Trip Types (rows 7-18, 0-indexed: 6-17)
    # Row 6: headers [Row Labels, D+I, I+I, 点点, Grand Total]
    triptype_data = []
    for r in range(7, 17):  # rows 8-17 (airlines)
        if r >= len(g) or not g[r][0]:
            continue
        airline = g[r][0]
        if airline == 'Grand Total':
            continue
        triptype_data.append({
            'airline': str(airline),
            'D+I': int(g[r][1] or 0),
            'I+I': int(g[r][2] or 0),
            'P2P': int(g[r][3] or 0),
            'total': int(g[r][4] or 0),
        })
    data['triptype'] = triptype_data
    # Grand total
    for r in range(7, 20):
        if r < len(g) and g[r][0] == 'Grand Total':
            data['triptype_total'] = {
                'D+I': int(g[r][1] or 0),
                'I+I': int(g[r][2] or 0),
                'P2P': int(g[r][3] or 0),
                'total': int(g[r][4] or 0),
            }
            break

    # Table 2: GREEN section (rows 26-35, 0-indexed: 25-34)
    # Row 25: headers
    # Cols F-L (5-11): SHARE by month MAY-OCT + TTL
    # Cols N-T (13-19): Current week JUNE-NOV + TTL (col 20)
    # Col V (21): Previous identifier
    # Col W (22): 总计环比
    # Cols X-AC (23-28): Previous week JUNE-NOV
    # Col AD (29): Previous TTL
    months_share = ['MAY', 'JUNE', 'JULY', 'AUG', 'SEP', 'OCT']
    months_current = ['JUNE', 'JULY', 'AUG', 'SEP', 'OCT', 'NOV']

    share_data = []
    for r in range(26, 35):  # rows 27-35
        if r >= len(g) or not g[r][5]:
            continue
        airline = str(g[r][5])
        if airline == 'TTL':
            airline = 'TOTAL'

        entry = {'airline': airline}

        # Monthly shares (cols 6-11)
        for mi, m in enumerate(months_share):
            val = g[r][6 + mi]
            entry[f'share_{m}'] = float(val) if val is not None else 0

        # Total share (col 12)
        entry['share_TTL'] = float(g[r][12]) if g[r][12] is not None else 0

        # Current week absolute (cols 14-19)
        for mi, m in enumerate(months_current):
            val = g[r][14 + mi]
            entry[f'curr_{m}'] = int(val) if val is not None and isinstance(val, (int, float)) else 0

        # Current total (col 20)
        entry['curr_TTL'] = int(g[r][20]) if g[r][20] is not None else 0

        # Previous total (col 29)
        prev_ttl = g[r][29] if len(g[r]) > 29 else None
        entry['prev_TTL'] = int(prev_ttl) if prev_ttl is not None else 0

        # WoW change (col 22)
        wow = g[r][22]
        entry['wow_pct'] = float(wow) if wow is not None and wow != '#DIV/0!' else 0

        # Previous week absolute (cols 23-28)
        for mi, m in enumerate(months_current):
            val = g[r][23 + mi] if len(g[r]) > 23 + mi else None
            entry[f'prev_{m}'] = int(val) if val is not None and isinstance(val, (int, float)) else 0

        # Monthly WoW changes (cols 30-35)
        for mi, m in enumerate(months_current):
            val = g[r][30 + mi] if len(g[r]) > 30 + mi else None
            if val is not None and val != '#DIV/0!':
                entry[f'wow_{m}'] = float(val)
            else:
                entry[f'wow_{m}'] = None

        # Previous share (col 37)
        prev_share = g[r][37] if len(g[r]) > 37 else None
        entry['prev_share'] = float(prev_share) if prev_share is not None else 0

        share_data.append(entry)

    data['share'] = share_data

    # Table 3: Bottom tables (rows 62-71, 0-indexed: 61-70)
    # Three sub-tables side by side:
    # Cols F-I (5-8): absolute D+I, I+I, P2P, TTL
    # Cols K-N (10-13): share within market (% of total per trip type)
    # Cols P-S (15-18): share within airline
    bottom_data = []
    for r in range(62, 71):
        if r >= len(g) or not g[r][5]:
            continue
        airline = str(g[r][5])
        if airline == 'Grand Total':
            airline = 'TOTAL'

        entry = {'airline': airline}

        # Absolute
        entry['abs_DI'] = int(g[r][6] or 0)
        entry['abs_II'] = int(g[r][7] or 0)
        entry['abs_P2P'] = int(g[r][8] or 0)
        entry['abs_TTL'] = int(g[r][9] or 0)

        # Share in market (% of column total)
        entry['mkt_DI'] = float(g[r][11]) if g[r][11] is not None else 0
        entry['mkt_II'] = float(g[r][12]) if g[r][12] is not None else 0
        entry['mkt_P2P'] = float(g[r][13]) if g[r][13] is not None else 0
        entry['mkt_TTL'] = float(g[r][14]) if g[r][14] is not None else 0

        # Share within airline (% of row total)
        entry['own_DI'] = float(g[r][16]) if g[r][16] is not None else 0
        entry['own_II'] = float(g[r][17]) if g[r][17] is not None else 0
        entry['own_P2P'] = float(g[r][18]) if g[r][18] is not None else 0

        bottom_data.append(entry)

    data['triptype_detail'] = bottom_data

    return data


# ─── Sheet 2: ALL AGTS ───

def parse_sheet2(wb):
    """Parse ALL AGTS sheet - agent × airline matrix."""
    ws = wb['ALL AGTS']
    g = read_sheet_as_grid(ws, max_row=200)

    # Row 8 (idx 8): headers [Row Labels, CA, CZ, MU, HU, HO, BA, ZH, JD, GS, NZ, Grand Total]
    # Find header row
    header_row = None
    for i, row in enumerate(g):
        if row and row[0] == 'Row Labels':
            header_row = i
            break

    if header_row is None:
        return []

    airlines = []
    for j in range(1, min(12, len(g[header_row]))):
        val = g[header_row][j]
        if val and val != 'Grand Total':
            airlines.append(str(val))

    agents = []
    for r in range(header_row + 1, len(g)):
        if not g[r] or not g[r][0]:
            continue
        name = str(g[r][0])
        if name == 'Grand Total':
            break

        entry = {'agent': name}
        total = 0
        for j, al in enumerate(airlines):
            val = g[r][1 + j]
            pax = int(val) if val is not None and isinstance(val, (int, float)) else 0
            entry[al] = pax
            total += pax

        gt = g[r][len(airlines) + 1]
        entry['total'] = int(gt) if gt is not None else total
        agents.append(entry)

    return {'airlines': airlines, 'agents': agents}


# ─── Sheet 3: TOP 30 OD - ALL ───

def parse_sheet3(wb):
    """Parse TOP 30 OD - ALL: CA top ODs + all airlines comparison."""
    ws = wb['TOP 30 OD - ALL']
    g = read_sheet_as_grid(ws, max_row=45)

    # Left table: CA top ODs (cols A-B, rows 9-38)
    ca_ods = []
    for r in range(8, 40):
        if r >= len(g) or not g[r][0] or g[r][0] == 'Grand Total':
            break
        ca_ods.append({'od': str(g[r][0]), 'pax': int(g[r][1] or 0)})

    # Right table: All airlines (cols E-P, rows 9-38)
    # Row 7 (idx 7): headers
    all_airlines = []
    if len(g) > 7:
        for j in range(5, 16):
            if j < len(g[7]) and g[7][j] and g[7][j] not in ('Row Labels', 'Grand Total'):
                all_airlines.append(str(g[7][j]))

    all_ods = []
    for r in range(8, 40):
        if r >= len(g) or not g[r][4] or g[r][4] == 'Grand Total':
            break
        entry = {'od': str(g[r][4])}
        for j, al in enumerate(all_airlines):
            val = g[r][5 + j]
            entry[al] = int(val) if val is not None and isinstance(val, (int, float)) else 0
        gt = g[r][5 + len(all_airlines)]
        entry['total'] = int(gt) if gt is not None else 0
        all_ods.append(entry)

    # ─── Aggregate to city pairs and country pairs ───
    AIRPORT_TO_CITY = {
        'LHR': 'London', 'LGW': 'London', 'STN': 'London', 'LTN': 'London',
        'MAN': 'Manchester', 'EDI': 'Edinburgh',
        'PEK': 'Beijing', 'PKX': 'Beijing',
        'PVG': 'Shanghai', 'SHA': 'Shanghai',
        'CAN': 'Guangzhou', 'SZX': 'Shenzhen',
        'CTU': 'Chengdu', 'TFU': 'Chengdu',
        'HGH': 'Hangzhou', 'NKG': 'Nanjing',
        'XIY': "Xi'an", 'WUH': 'Wuhan',
        'CGO': 'Zhengzhou', 'CSX': 'Changsha',
        'TAO': 'Qingdao', 'FOC': 'Fuzhou',
        'SHE': 'Shenyang', 'KMG': 'Kunming',
        'CKG': 'Chongqing',
        'HKG': 'Hong Kong', 'NRT': 'Tokyo', 'HND': 'Tokyo',
        'KIX': 'Osaka', 'ICN': 'Seoul',
        'BKK': 'Bangkok', 'MNL': 'Manila',
        'SIN': 'Singapore', 'KUL': 'Kuala Lumpur',
        'SYD': 'Sydney', 'MEL': 'Melbourne', 'AKL': 'Auckland',
        'TPE': 'Taipei', 'SGN': 'Ho Chi Minh',
        'CNX': 'Chiang Mai', 'UBN': 'Ulaanbaatar',
    }

    CITY_TO_COUNTRY = {
        'London': 'UK', 'Manchester': 'UK', 'Edinburgh': 'UK',
        'Beijing': 'China', 'Shanghai': 'China', 'Guangzhou': 'China',
        'Shenzhen': 'China', 'Chengdu': 'China', 'Hangzhou': 'China',
        'Nanjing': 'China', "Xi'an": 'China', 'Wuhan': 'China',
        'Zhengzhou': 'China', 'Changsha': 'China', 'Qingdao': 'China',
        'Fuzhou': 'China', 'Shenyang': 'China', 'Kunming': 'China',
        'Chongqing': 'China',
        'Hong Kong': 'HK/Other', 'Tokyo': 'Japan', 'Osaka': 'Japan',
        'Seoul': 'Korea', 'Bangkok': 'Thailand', 'Manila': 'Philippines',
        'Singapore': 'Singapore', 'Kuala Lumpur': 'Malaysia',
        'Sydney': 'Australia', 'Melbourne': 'Australia', 'Auckland': 'New Zealand',
        'Taipei': 'Taiwan', 'Ho Chi Minh': 'Vietnam',
        'Chiang Mai': 'Thailand', 'Ulaanbaatar': 'Mongolia',
    }

    def get_city(apt):
        return AIRPORT_TO_CITY.get(apt, apt)

    def make_pair_key(orig, dest):
        """UK is always origin in this sheet, so keep direction."""
        return f'{orig} → {dest}'

    # City pair aggregation
    city_pair_map = {}
    for od in all_ods:
        code = od['od']
        orig_apt, dest_apt = code[:3], code[3:]
        orig_city, dest_city = get_city(orig_apt), get_city(dest_apt)
        cpkey = make_pair_key(orig_city, dest_city)

        if cpkey not in city_pair_map:
            city_pair_map[cpkey] = {al: 0 for al in all_airlines}
            city_pair_map[cpkey]['total'] = 0
            city_pair_map[cpkey]['city_pair'] = cpkey
            city_pair_map[cpkey]['airports'] = []

        for al in all_airlines:
            city_pair_map[cpkey][al] += od.get(al, 0)
        city_pair_map[cpkey]['total'] += od.get('total', 0)
        city_pair_map[cpkey]['airports'].append(code)

    city_pairs = sorted(city_pair_map.values(), key=lambda x: x['total'], reverse=True)

    # Country pair aggregation
    country_pair_map = {}
    for cp in city_pairs:
        pair = cp['city_pair']
        cities = pair.split(' → ')
        if len(cities) == 2:
            c1 = CITY_TO_COUNTRY.get(cities[0], 'Other')
            c2 = CITY_TO_COUNTRY.get(cities[1], 'Other')
            cpkey = make_pair_key(c1, c2)
        else:
            cpkey = 'Other'

        if cpkey not in country_pair_map:
            country_pair_map[cpkey] = {al: 0 for al in all_airlines}
            country_pair_map[cpkey]['total'] = 0
            country_pair_map[cpkey]['country_pair'] = cpkey

        for al in all_airlines:
            country_pair_map[cpkey][al] += cp.get(al, 0)
        country_pair_map[cpkey]['total'] += cp.get('total', 0)

    country_pairs = sorted(country_pair_map.values(), key=lambda x: x['total'], reverse=True)

    return {
        'ca_ods': ca_ods,
        'all_airlines': all_airlines,
        'all_ods': all_ods,
        'city_pairs': city_pairs,
        'country_pairs': country_pairs,
    }


# ─── Sheet 4: CONSOL ───

def parse_sheet4(wb):
    """Parse CONSOL sheet - dynamically find sections."""
    ws = wb['CONSOL']
    g = read_sheet_as_grid(ws, max_row=65, max_col=50)

    sections = find_sections(g, marker_col_start=10, marker_col_end=30)

    share_data = []
    present_data = []
    weekly_comparison = []

    for sec in sections:
        if not sec['airlines']:
            continue
        if sec['type'] == 'share' or (sec['type'] == 'unknown' and not share_data):
            share_data = read_section_data(g, sec, is_pct=True)
        elif sec['type'] in ('present',) and not present_data:
            present_data = read_section_data(g, sec, is_pct=False)
        elif sec['type'] == 'weekly' and not weekly_comparison:
            # Weekly with curr + last side by side
            r_start = sec['row'] + 1
            col = sec['col']
            als = sec['airlines']
            total_col = col + 1 + len(als)
            last_start = total_col + 2  # skip Total col + agent name repeat
            wc = []
            for r in range(r_start, min(r_start + 20, len(g))):
                if r >= len(g) or not g[r] or not g[r][col]: continue
                agent = str(g[r][col])
                if agent in ('Agents', 'Weekly Comparison'): continue
                if agent in ('Grand Total', 'Total'): agent = 'TOTAL'
                nv = g[r][col + 1] if col + 1 < len(g[r] or []) else None
                if isinstance(nv, str): continue
                entry = {'agent': agent}
                for j, al in enumerate(als):
                    val = g[r][col + 1 + j] if col + 1 + j < len(g[r] or []) else None
                    entry[f'{al}_curr'] = int(val) if val is not None and isinstance(val, (int, float)) else 0
                t_val = g[r][total_col] if total_col < len(g[r] or []) else None
                entry['total_curr'] = int(t_val) if t_val is not None and isinstance(t_val, (int, float)) else 0
                for j, al in enumerate(als):
                    val = g[r][last_start + j] if last_start + j < len(g[r] or []) else None
                    entry[f'{al}_last'] = int(val) if val is not None and isinstance(val, (int, float)) else 0
                wc.append(entry)
                if entry.get("agent") == "TOTAL" or entry.get("tmc") == "TOTAL":
                    break
            weekly_comparison = wc

    # Airline summary: search for triptype or YTD table in bottom area
    airline_summary = []
    for r in range(35, min(65, len(g))):
        if r >= len(g) or not g[r]: continue
        # Search in col 15 area for "Row Labels"
        for c in range(10, 20):
            if c < len(g[r]) and g[r][c] == 'Row Labels':
                next_col = g[r][c + 1] if c + 1 < len(g[r]) else None
                if next_col in ('D+I', 'I+I'):
                    for rr in range(r + 1, min(r + 15, len(g))):
                        if rr >= len(g) or not g[rr][c]: continue
                        al = str(g[rr][c])
                        if al in ('TTL', 'Grand Total'): break
                        entry = {'airline': al}
                        for j, cn in enumerate(['D+I', 'I+I', 'P2P', 'total']):
                            val = g[rr][c + 1 + j] if c + 1 + j < len(g[rr]) else None
                            entry[cn] = int(val) if val is not None and isinstance(val, (int, float)) else 0
                        airline_summary.append(entry)
                break
        if airline_summary:
            break

    return {
        'share': share_data,
        'present': present_data,
        'weekly_comparison': weekly_comparison,
        'airline_summary': airline_summary,
    }


# ─── Helper: find table sections by header markers ───

def find_sections(g, marker_col_start=20, marker_col_end=45):
    """Scan grid for 'Agents' header rows in the right-side area.
    Returns list of {'type': ..., 'row': header_row, 'col': agent_col, 'airlines': [...]}
    """
    sections = []
    for r in range(len(g)):
        for c in range(marker_col_start, min(marker_col_end, len(g[r]) if g[r] else 0)):
            val = g[r][c]
            if val == 'Agents':
                # Read airline names from this header row
                airlines = []
                for j in range(c + 1, min(c + 15, len(g[r]))):
                    h = g[r][j]
                    if h is None or h in ('Grand Total', 'Total'):
                        break
                    if isinstance(h, str) and len(h) <= 3 and h.isalpha():
                        airlines.append(h)
                # Determine section type from row above
                section_type = 'unknown'
                if r > 0:
                    for cc in range(c, min(c + 5, len(g[r-1]) if g[r-1] else 0)):
                        prev = g[r-1][cc] if g[r-1] else None
                        if prev and isinstance(prev, str):
                            pv = prev.strip().lower()
                            if 'present week' in pv or 'total' == pv:
                                section_type = 'present'
                            elif 'weekly comparison' in pv:
                                section_type = 'weekly'
                            elif 'weekly share' in pv:
                                section_type = 'weekly_share'
                            elif 'previous' in pv or 'past' in pv:
                                section_type = 'previous_share'
                            break
                # If still unknown, check if values in next row are floats (share) or ints (pax)
                if section_type == 'unknown' and r + 1 < len(g):
                    sample = g[r+1][c+1] if c+1 < len(g[r+1] or []) else None
                    if isinstance(sample, float) and 0 < sample < 1:
                        section_type = 'share'
                    elif isinstance(sample, (int, float)) and sample > 1:
                        section_type = 'present'
                sections.append({'type': section_type, 'row': r, 'col': c, 'airlines': airlines})
    return sections


def read_section_data(g, section, is_pct=False):
    """Read rows after a section header until Total/Grand Total row (inclusive)."""
    data = []
    r_start = section['row'] + 1
    col = section['col']
    airlines = section['airlines']
    total_col = col + 1 + len(airlines)  # col after last airline

    for r in range(r_start, min(r_start + 25, len(g))):
        if r >= len(g) or not g[r] or not g[r][col]:
            continue
        agent = str(g[r][col])
        # Skip known section markers
        if agent in ('Agents', 'Row Labels', 'Weekly Comparison', 'Weekly Share',
                      'Current Week Share', 'Past Week Share', 'Growth weekly',
                      'Previous Week Share', 'PRESENT WEEK', 'PREVIOUS'):
            continue
        is_total = agent in ('Total', 'Grand Total', 'TOTAL')
        if is_total:
            agent = 'TOTAL'

        # Skip if next cell is text (sub-header)
        next_val = g[r][col + 1] if col + 1 < len(g[r] or []) else None
        if isinstance(next_val, str) and next_val in airlines:
            continue

        entry = {'agent': agent}
        for j, al in enumerate(airlines):
            val = g[r][col + 1 + j] if col + 1 + j < len(g[r] or []) else None
            if is_pct:
                entry[al] = float(val) * 100 if val is not None and isinstance(val, (int, float)) else 0
            else:
                entry[al] = int(val) if val is not None and isinstance(val, (int, float)) else 0

        # Total column
        t_val = g[r][total_col] if total_col < len(g[r] or []) else None
        if not is_pct:
            entry['total'] = int(t_val) if t_val is not None and isinstance(t_val, (int, float)) else sum(entry.get(al, 0) for al in airlines)

        data.append(entry)
        if is_total:
            break  # Stop after Total row, don't read into next section
    return data


# ─── Sheet 6: OTA ───

def parse_sheet6(wb):
    """Parse OTA sheet - dynamically find sections by header markers."""
    ws = wb['OTA']
    g = read_sheet_as_grid(ws, max_row=85, max_col=55)

    sections = find_sections(g, marker_col_start=20, marker_col_end=45)

    share_data = []
    present_data = []
    weekly_comparison = []
    weekly_share_curr = []
    ota_airlines = []

    for sec in sections:
        if not sec['airlines']:
            continue
        if not ota_airlines:
            ota_airlines = sec['airlines']

        if sec['type'] == 'share' or (sec['type'] == 'unknown' and not share_data):
            share_data = read_section_data(g, sec, is_pct=True)
        elif sec['type'] == 'present' and not present_data:
            present_data = read_section_data(g, sec, is_pct=False)
        elif sec['type'] == 'weekly':
            # Weekly comparison: two sets of airlines side by side (curr + last)
            raw = read_section_data(g, sec, is_pct=False)
            # The section has airlines for current, then after Total col, same airlines for last week
            # Re-read with extended columns
            wc = []
            r_start = sec['row'] + 1
            col = sec['col']
            als = sec['airlines']
            total_col = col + 1 + len(als)
            last_start = total_col + 2  # skip Total col + agent name repeat  # skip Total, then last week airlines

            for r in range(r_start, min(r_start + 20, len(g))):
                if r >= len(g) or not g[r] or not g[r][col]:
                    continue
                agent = str(g[r][col])
                if agent in ('Agents', 'Weekly Comparison'):
                    continue
                if agent in ('Total', 'Grand Total'):
                    agent = 'TOTAL'
                next_val = g[r][col + 1] if col + 1 < len(g[r] or []) else None
                if isinstance(next_val, str):
                    continue

                entry = {'agent': agent}
                for j, al in enumerate(als):
                    val = g[r][col + 1 + j] if col + 1 + j < len(g[r] or []) else None
                    entry[f'{al}_curr'] = int(val) if val is not None and isinstance(val, (int, float)) else 0
                t_val = g[r][total_col] if total_col < len(g[r] or []) else None
                entry['total_curr'] = int(t_val) if t_val is not None and isinstance(t_val, (int, float)) else 0
                # Last week values
                for j, al in enumerate(als):
                    val = g[r][last_start + j] if last_start + j < len(g[r] or []) else None
                    entry[f'{al}_last'] = int(val) if val is not None and isinstance(val, (int, float)) else 0
                wc.append(entry)
                if entry.get("agent") == "TOTAL" or entry.get("tmc") == "TOTAL":
                    break
            weekly_comparison = wc
        elif sec['type'] == 'weekly_share':
            weekly_share_curr = read_section_data(g, sec, is_pct=True)

    return {
        'airlines': ota_airlines[:6] if ota_airlines else ['CA', 'MU', 'CZ', 'HU', 'ZH', 'HO'],
        'share': share_data,
        'present': present_data,
        'weekly_comparison': weekly_comparison,
        'weekly_share': weekly_share_curr,
    }


# ─── Sheet 7: TMC ───

def parse_sheet7(wb):
    """Parse TMC sheet - dynamically find sections."""
    ws = wb['TMC']
    g = read_sheet_as_grid(ws, max_row=75, max_col=45)

    # TMC uses named TMC companies instead of "Agents" as marker.
    # Search for rows with TMC name pattern: row has a known TMC name at some col,
    # and the row above has airline codes (CA, BA, MU...).
    # Use find_sections but also search for "PRESENT WEEK" / TMC names

    # Strategy: find header rows that contain airline codes like CA, BA, MU
    def find_tmc_sections():
        sections = []
        for r in range(len(g)):
            for c in range(8, min(30, len(g[r]) if g[r] else 0)):
                val = g[r][c]
                # Look for known TMC names or "PRESENT WEEK" marker
                if val and isinstance(val, str) and val in ('PRESENT WEEK', 'PREVIOUS', 'CURRENT WEEK'):
                    # Next row should have TMC names and airline values
                    if r + 1 < len(g) and g[r + 1]:
                        tmc_name = g[r + 1][c]
                        if tmc_name and isinstance(tmc_name, str):
                            # Read airline headers from the row with this TMC
                            # The airlines are in the header one row before data
                            airlines = []
                            # Check if same row has airline names after TMC col
                            for j in range(c + 1, min(c + 12, len(g[r + 1]) if g[r + 1] else 0)):
                                h = g[r][j] if g[r] else None  # header row = marker row
                                # Actually check: the marker row might have "CA", "BA", etc.
                                pass

                            # Better approach: find a row that starts with TMC names and has numbers
                            # Look for the Agents-like header
                            pass
                    sections.append({'marker': val, 'row': r, 'col': c})
        return sections

    # Even simpler: use the same find_sections but search wider and for TMC-specific patterns
    # TMC section headers use TMC company names directly, with airline cols as values
    # Let me just scan for rows where col has a known TMC name and col+1 has a number

    sections = find_sections(g, marker_col_start=8, marker_col_end=25)

    # Also scan for "PRESENT WEEK" / "PREVIOUS" markers
    present_markers = []
    for r in range(len(g)):
        for c in range(8, min(25, len(g[r]) if g[r] else 0)):
            val = g[r][c]
            if val and isinstance(val, str):
                vl = val.strip()
                if vl in ('PRESENT WEEK', 'CURRENT WEEK'):
                    present_markers.append({'type': 'present_header', 'row': r, 'col': c})
                elif vl == 'PREVIOUS':
                    present_markers.append({'type': 'previous_header', 'row': r, 'col': c})

    present_data = []
    share_data = []
    weekly_comparison = []
    tmc_airlines = []

    # Use find_sections results if available
    for sec in sections:
        if not sec['airlines']:
            continue
        if not tmc_airlines:
            tmc_airlines = sec['airlines']

        # Rename 'agent' key to 'tmc' in results
        if sec['type'] == 'share' or (sec['type'] == 'unknown' and not share_data):
            raw = read_section_data(g, sec, is_pct=True)
            share_data = [{'tmc': d.pop('agent'), **d} for d in raw]
        elif sec['type'] == 'present' and not present_data:
            raw = read_section_data(g, sec, is_pct=False)
            present_data = [{'tmc': d.pop('agent'), **d} for d in raw]
        elif sec['type'] == 'weekly' and not weekly_comparison:
            r_start = sec['row'] + 1
            col = sec['col']
            als = sec['airlines']
            total_col = col + 1 + len(als)
            last_start = total_col + 2  # skip Total col + agent name repeat
            wc = []
            for r in range(r_start, min(r_start + 20, len(g))):
                if r >= len(g) or not g[r] or not g[r][col]: continue
                tmc = str(g[r][col])
                if tmc in ('Agents', 'Weekly Comparison', 'PRESENT WEEK', 'PREVIOUS'): continue
                if tmc in ('Grand Total', 'Total'): tmc = 'TOTAL'
                nv = g[r][col + 1] if col + 1 < len(g[r] or []) else None
                if isinstance(nv, str): continue
                entry = {'tmc': tmc}
                for j, al in enumerate(als):
                    val = g[r][col + 1 + j] if col + 1 + j < len(g[r] or []) else None
                    entry[al] = int(val) if val is not None and isinstance(val, (int, float)) else 0
                t_val = g[r][total_col] if total_col < len(g[r] or []) else None
                entry['total'] = int(t_val) if t_val is not None and isinstance(t_val, (int, float)) else 0
                wc.append(entry)
                if entry.get("agent") == "TOTAL" or entry.get("tmc") == "TOTAL":
                    break
            weekly_comparison = wc

    # If find_sections didn't find TMC data (different structure),
    # try present_markers approach
    if not present_data and present_markers:
        for pm in present_markers:
            if pm['type'] == 'present_header':
                r = pm['row']
                c = pm['col']
                # The TMC names are at col c, airlines at col c+1 onwards
                # Read row r to find TMC name and next row for data
                # Actually: marker row has TMC names listed vertically below
                # and airline values horizontally
                # Read the next rows as TMC data
                als = []
                # Find airline header - usually same row as PRESENT WEEK or one below
                for rr in range(r, min(r + 2, len(g))):
                    for j in range(c + 1, min(c + 12, len(g[rr]) if g[rr] else 0)):
                        h = g[rr][j]
                        if h and isinstance(h, str) and len(h) <= 3 and h.isalpha() and h.isupper():
                            als.append(h)
                    if als:
                        break

                if not als:
                    continue
                if not tmc_airlines:
                    tmc_airlines = als

                # Read TMC rows: col c = TMC name, col c+1..c+n = airline values
                for rr in range(r + 1, min(r + 25, len(g))):
                    if rr >= len(g) or not g[rr] or not g[rr][c]: continue
                    tmc = str(g[rr][c])
                    if tmc in ('PRESENT WEEK', 'PREVIOUS', 'CURRENT WEEK'): continue
                    if tmc == 'Grand Total':
                        tmc = 'TOTAL'
                    nv = g[rr][c + 1] if c + 1 < len(g[rr] or []) else None
                    if isinstance(nv, str) and nv in als:
                        continue  # another header row
                    entry = {'tmc': tmc}
                    for j, al in enumerate(als):
                        val = g[rr][c + 1 + j] if c + 1 + j < len(g[rr] or []) else None
                        entry[al] = int(val) if val is not None and isinstance(val, (int, float)) else 0
                    # Total
                    tc = c + 1 + len(als)
                    t_val = g[rr][tc] if tc < len(g[rr] or []) else None
                    entry['total'] = int(t_val) if t_val is not None and isinstance(t_val, (int, float)) else sum(entry.get(al, 0) for al in als)
                    present_data.append(entry)
                break

    return {
        'airlines': tmc_airlines if tmc_airlines else ['CA', 'BA', 'MU', 'CZ', 'HU', 'ZH', 'HO', 'GS'],
        'present': present_data,
        'share': share_data,
        'weekly_comparison': weekly_comparison,
    }


# ─── Sheet 8: Trip.com ───

def parse_sheet8(wb):
    """Parse Trip.com sheet - by POO country."""
    ws = wb['Trip.com']
    g = read_sheet_as_grid(ws, max_row=40)

    # Row 5 (idx 5): headers [Row Labels, CA, MU, CZ, HU, HO, ZH, GS, JD, BA, Grand Total]
    trip_airlines = ['CA', 'MU', 'CZ', 'HU', 'HO', 'ZH', 'GS', 'JD', 'BA']

    countries = []
    for r in range(6, 35):
        if r >= len(g) or not g[r][0]:
            continue
        country = str(g[r][0])
        if country == 'Grand Total':
            country = 'TOTAL'
        entry = {'country': country}
        for j, al in enumerate(trip_airlines):
            val = g[r][1 + j]
            entry[al] = int(val) if val is not None and isinstance(val, (int, float)) else 0
        gt = g[r][10] if len(g[r]) > 10 else None
        entry['total'] = int(gt) if gt is not None else 0
        countries.append(entry)

    return {'airlines': trip_airlines, 'countries': countries}


# ─── Raw grid export ───

RAW_SHEET_MAP = {
    'Month share - Pax': 'raw_month_share',
    'ALL AGTS':          'raw_all_agts',
    'TOP 30 OD - ALL':   'raw_top30_od_all',
    'CONSOL':            'raw_consol',
    'TOP 30 OD - OTA':   'raw_top30_od_ota',
    'OTA':               'raw_ota',
    'TMC':               'raw_tmc',
    'Trip.com':          'raw_trip_com',
}


def _serialize_cell(v):
    """Convert a cell value to something JSON-safe."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    return str(v)


def parse_raw_grids(wb):
    """Read each Excel sheet into a raw 2D array (up to 200 rows x 50 cols).
    Returns dict mapping safe_name -> grid (list of lists).
    """
    result = {}
    for sheet_name, safe_name in RAW_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"   [raw] Sheet '{sheet_name}' not found, skipping", flush=True)
            continue
        ws = wb[sheet_name]
        grid = read_sheet_as_grid(ws, max_row=200, max_col=50)
        # Convert all cells to JSON-safe values
        grid = [[_serialize_cell(c) for c in row] for row in grid]
        # Trim trailing all-None rows
        while grid and all(c is None for c in grid[-1]):
            grid.pop()
        # Trim trailing None cols from each row
        if grid:
            max_used_col = 0
            for row in grid:
                for ci in range(len(row) - 1, -1, -1):
                    if row[ci] is not None:
                        max_used_col = max(max_used_col, ci + 1)
                        break
            grid = [row[:max_used_col] for row in grid]
        result[safe_name] = grid
        print(f"   [raw] {safe_name}: {len(grid)} rows x {max_used_col if grid else 0} cols", flush=True)
    return result


# ─── Main ───

def parse_sheets(filepath, db_path=None):
    if db_path is None:
        db_path = DB_PATH

    filepath = os.path.expanduser(filepath)
    snapshot = extract_snapshot_date(filepath)

    print(f"[1/3] Opening workbook (data_only)...", flush=True)
    t0 = time.time()
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    print(f"   Opened in {time.time()-t0:.1f}s", flush=True)

    print(f"[2/3] Parsing sheets...", flush=True)

    sheet1 = parse_sheet1(wb)
    sheet2 = parse_sheet2(wb)
    sheet3 = parse_sheet3(wb)
    sheet4 = parse_sheet4(wb)
    sheet6 = parse_sheet6(wb)
    sheet7 = parse_sheet7(wb)
    sheet8 = parse_sheet8(wb)

    print(f"   Parsing raw grids...", flush=True)
    raw_grids = parse_raw_grids(wb)

    wb.close()
    print(f"   All sheets parsed in {time.time()-t0:.1f}s", flush=True)

    # Store as JSON in SQLite
    print(f"[3/3] Storing in SQLite...", flush=True)
    db_path_abs = os.path.abspath(db_path)
    os.makedirs(os.path.dirname(db_path_abs), exist_ok=True)
    conn = sqlite3.connect(db_path_abs)
    c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS sheet_data (
        snapshot_date TEXT,
        sheet_name TEXT,
        data_json TEXT,
        PRIMARY KEY (snapshot_date, sheet_name)
    )""")

    sheets = {
        'month_share': sheet1,
        'all_agents': sheet2,
        'top30_od': sheet3,
        'consol': sheet4,
        'ota': sheet6,
        'tmc': sheet7,
        'trip_com': sheet8,
    }

    for name, data in sheets.items():
        c.execute("""INSERT OR REPLACE INTO sheet_data VALUES (?, ?, ?)""",
                  (snapshot, name, json.dumps(data, ensure_ascii=False)))

    # Store raw grids
    for safe_name, grid in raw_grids.items():
        c.execute("""INSERT OR REPLACE INTO sheet_data VALUES (?, ?, ?)""",
                  (snapshot, safe_name, json.dumps(grid, ensure_ascii=False)))

    conn.commit()
    conn.close()

    print(f"   Done! Snapshot: {snapshot}", flush=True)
    print(f"   Sheets stored: {list(sheets.keys())}", flush=True)
    print(f"   Total time: {time.time()-t0:.1f}s", flush=True)

    # Print summary
    print(f"\n   Sheet1 share data: {len(sheet1.get('share', []))} airlines", flush=True)
    print(f"   Sheet2 agents: {len(sheet2.get('agents', []))}", flush=True)
    print(f"   Sheet3 ODs: CA={len(sheet3.get('ca_ods', []))}, ALL={len(sheet3.get('all_ods', []))}", flush=True)
    print(f"   Sheet4 CONSOL: {len(sheet4.get('present', []))} agents", flush=True)
    print(f"   Sheet6 OTA: {len(sheet6.get('present', []))} agents", flush=True)
    print(f"   Sheet7 TMC: {len(sheet7.get('present', []))} TMCs", flush=True)
    print(f"   Sheet8 Trip.com: {len(sheet8.get('countries', []))} countries", flush=True)

    return snapshot


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python sheet_parser.py <excel_file> [--db DB_PATH]")
        sys.exit(1)
    db_path = None
    if '--db' in sys.argv:
        idx = sys.argv.index('--db')
        if idx + 1 < len(sys.argv):
            db_path = sys.argv[idx + 1]
    if not db_path:
        db_path = os.environ.get('DB_PATH', DB_PATH)
    parse_sheets(sys.argv[1], db_path=db_path)
